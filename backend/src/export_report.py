from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# (sheet title, results key, columns to keep in that order)
PHASES = [
    ("Meter Coverage", "step0", ["Match Key", "Value", "Client Name", "Issue"]),
    ("Metadata", "step1", ["Meter Number", "Client Name", "Mismatched Field", "Original Field (Hebrew)", "Alteco Value", "Client Value"]),
    ("Consumption", "step2", ["Meter Number", "Client Name", "Mismatched Field", "Original Field (Hebrew)", "Alteco Value", "Client Value"]),
    ("Financial", "step3", ["Customer ID", "Client Name", "Mismatched Field", "Original Field (Hebrew)", "Alteco Value", "Client Value"]),
]

HEADER_FILL = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def _style_sheet(ws, ncols):
    for col_idx in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    for col_idx in range(1, ncols + 1):
        letter = get_column_letter(col_idx)
        max_len = max(
            (len(str(ws.cell(row=r, column=col_idx).value or "")) for r in range(1, ws.max_row + 1)),
            default=10,
        )
        ws.column_dimensions[letter].width = min(max(max_len + 2, 12), 60)


def build_discrepancy_workbook(results):
    """
    Builds a formatted .xlsx report from a reconciliation results dict
    ({"step0": [...], "step1": [...], "step2": [...], "step3": [...]}).
    Sheet order: Summary first, then one sheet per phase (bold header row,
    frozen header, auto-filter, auto-sized columns) — empty phases still get
    a sheet, with a clear "no discrepancies" note instead of being skipped.
    Returns a BytesIO positioned at 0, ready to stream as a response body.
    """
    buffer = BytesIO()
    with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
        summary_rows = [
            {"Phase": title, "Mismatches": len(results.get(key) or [])}
            for title, key, _cols in PHASES
        ]
        df_summary = pd.DataFrame(summary_rows)
        df_summary.to_excel(writer, sheet_name="Summary", index=False)
        writer.sheets["Summary"]["C1"] = "Generated"
        writer.sheets["Summary"]["C2"] = datetime.now().strftime("%Y-%m-%d %H:%M")
        _style_sheet(writer.sheets["Summary"], len(df_summary.columns))

        for title, key, columns in PHASES:
            rows = results.get(key) or []
            if rows:
                df = pd.DataFrame(rows)
                df = df[[c for c in columns if c in df.columns]]
            else:
                df = pd.DataFrame([{"Result": "No discrepancies found"}])
            df.to_excel(writer, sheet_name=title, index=False)
            _style_sheet(writer.sheets[title], len(df.columns))

    buffer.seek(0)
    return buffer
